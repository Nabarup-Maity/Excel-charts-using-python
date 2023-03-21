#!/usr/bin/env python
# coding: utf-8

#1. Import the necessary packages:
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties


# 2. create an input dataframe
df = pd.DataFrame({
    'Time': ['2021-01-01', '2021-01-02', '2021-01-03', '2021-01-04', '2021-01-05','2021-01-06','2021-01-07'],
    'Value': [10, 15, 12, 17, 22, 14, 11],
    'Anomaly': [False, True, True, False, False, True, False]
})
df


# 3. Create a new workbook and select the active worksheet:
wb = Workbook()
ws = wb.active


# 4. write the pandas dataframe to the existing sheet ws with openpyxl
rows = dataframe_to_rows(df,index=False)
for row_idx, row in enumerate(rows,1):
    for c_idx, value in enumerate(row,1):
        ws.cell(row = row_idx, column = c_idx, value = value)



# 5. Create first chart c1: Bar chart where x axis = Time and y axis = Value
c1 = BarChart()
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.x_axis.title = "Date"
c1.y_axis.title = "Value"
c1.y_axis.majorGridlines = None
c1.title = 'Trend'

# Y Axis value
data = Reference(ws, min_col=2, min_row=1, max_row=df.shape[0]+1)
c1.add_data(data, titles_from_data=True)

# X axis value
dates = Reference(ws, min_col=1, min_row=2,  max_row=df.shape[0]+1)
c1.set_categories(dates)


# 6. Create a second chart c2: Line chart where x axis = Time and y axis = Value
c2 = LineChart()
c2.add_data(data, titles_from_data=True)
c2.y_axis.axId = 200


# 7. Update the line chart with color, size and shape
s1 = c2.series[0]
#choose symbol between{‘diamond’, ‘picture’, ‘star’, ‘plus’, ‘square’, ‘dash’, ‘dot’, ‘circle’, ‘x’, ‘auto’, ‘triangle’}
s1.marker.symbol = "circle"
s1.marker.size = 7
s1.marker.graphicalProperties.solidFill = "FFFF00" # Marker filling
s1.marker.graphicalProperties.line.solidFill = "FFFF00" # Marker outline


# 8. Set line color to red for anomalous values in the bar chart
s2 = c1.series[0]
for idx, val in enumerate(df['Value']):
    if df['Anomaly'][idx] == True:
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = 'FF0000'
        s2.dPt.append(pt)

# 9. Merge two graph into one
# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
c1.y_axis.crosses = "max"
c1 += c2

ws.add_chart(c1, "D4")


# 10. Save the workbook

wb.save("output.xlsx")


# ### Reference
# https://openpyxl.readthedocs.io/en/latest/charts/introduction.html





